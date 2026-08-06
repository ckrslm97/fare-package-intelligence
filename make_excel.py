"""Build a polished, formatted Excel from a run's outputs (raw_data.jsonl for
amenity details + normalized_data.csv for per-cabin source).

Columns: identity + route + brand + price, then one column per right
(Turkish label) showing "State — detail", colour-coded, plus miles + meta.

Usage: python make_excel.py <out_dir> [output.xlsx]
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from branded_fare_scraper.airports import meta as airport_meta
from branded_fare_scraper.amenities import AMENITY_KEYS, AMENITY_TR, canonical_rule_detail
from branded_fare_scraper.models import AmenityStatus
from branded_fare_scraper.normalization import (clean_fare_code, iter_unit_ranked_by_cabin,
                                                tier_code)
from branded_fare_scraper.rebuild import (cabin_result_from_dict, iter_raw_records,
                                          pair_prefs_for, season_pair_prefs)
from branded_fare_scraper.report import CARRIER_NAMES

_RANK = {AmenityStatus.INCLUDED: 3, AmenityStatus.PAID: 2, AmenityStatus.NOT_INCLUDED: 1, AmenityStatus.UNKNOWN: 0}
STATE_FILL = {
    "Included": PatternFill("solid", fgColor="C6EFCE"),
    "Paid": PatternFill("solid", fgColor="FFEB9C"),
    "Not Included": PatternFill("solid", fgColor="FFC7CE"),
    "Unknown": PatternFill("solid", fgColor="F2F2F2"),
}
STATE_FONT = {
    "Included": Font(color="1B6B3A"), "Paid": Font(color="9C6500"),
    "Not Included": Font(color="9C2A2A"), "Unknown": Font(color="A6A6A6"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")

BASE_COLS = ["Carrier", "Airline", "OND", "Origin", "Destination",
             "Origin City", "Origin Country", "Dest City", "Dest Country", "Season",
             "Departure", "Return", "Cabin", "Tier", "Brand Order",
             "Raw Brand Name", "Normalized Brand", "Fare Family Code",
             "Display Price", "Abs Price", "Currency", "Source"]
AMEN_COLS = [AMENITY_TR[k] for k in AMENITY_KEYS]
TAIL_COLS = ["Mil Var", "Mil", "Bonus %", "Scrape Time", "Status"]


def used_amenity_keys(amaps) -> list[str]:
    """Amenity keys some row actually reports (state not Unknown/blank).

    User rule: "çok fazla boş içerik duruyor" — a right that is Unknown on every
    single row is a dead column, so it is left out of the workbook. Raw data
    keeps the full taxonomy; only this display drops it.
    """
    used = {k for amap in amaps for k in AMENITY_KEYS
            if amap.get(k, ("Unknown", ""))[0] not in ("", "Unknown")}
    return [k for k in AMENITY_KEYS if k in used]     # taxonomy order


def main():
    out_dir = Path(sys.argv[1] if len(sys.argv) > 1 else "output_final")
    out_xlsx = Path(sys.argv[2]) if len(sys.argv) > 2 else out_dir / "branded_fares_formatted.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Branded Fares"

    scrape_time = datetime.now().replace(microsecond=0).isoformat()
    # Two passes: collect every row, then keep only the rights with real data.
    collected: list[tuple[list, dict, list]] = []
    raw_path = out_dir / "raw_data.jsonl"
    prefs = season_pair_prefs(raw_path)      # cross-season ladder-order consensus
    for rec in iter_raw_records(raw_path):
        ppc = pair_prefs_for(prefs, rec["carrier"], rec["origin"], rec["destination"])
        cabins = [cabin_result_from_dict(c) for c in rec.get("cabins", [])]
        # One ladder per effective cabin for the unit (no duplicate PE rows).
        for eff_cab, raw, nb, order, absp, c in iter_unit_ranked_by_cabin(
                cabins, carrier=rec["carrier"], pair_prefs_by_cabin=ppc):
            dep = c.departure.isoformat() if c.departure else None
            ret = c.return_date.isoformat() if c.return_date else None
            src = c.source or rec.get("source", "")   # per-cabin source from raw
            amap = {k: ("Unknown", "") for k in AMENITY_KEYS}
            for a in raw.amenities:
                k = a.canonical_key
                if k not in amap:
                    continue
                if _RANK[a.status] >= _RANK[AmenityStatus(amap[k][0])]:
                    amap[k] = (a.status.value, a.raw_value or "")   # tuple replace (no stale detail)
            mo = airport_meta(rec["origin"]); md = airport_meta(rec["destination"])
            row = [
                rec["carrier"], CARRIER_NAMES.get(rec["carrier"].upper(), rec["carrier"]),
                f'{rec["origin"]}-{rec["destination"]}', rec["origin"], rec["destination"],
                mo["city_code"], mo["country_code"], md["city_code"], md["country_code"],
                rec["season"], dep, ret, eff_cab.value, tier_code(eff_cab, order), order + 1,
                raw.raw_brand_name, nb.normalized_name, clean_fare_code(raw.fare_family_code),
                raw.display_price_text or "", (absp if absp is not None else raw.price_value),
                raw.currency or "USD", src,
            ]
            tail = [
                ("Evet" if raw.miles.mileage_available else "Hayır") if raw.miles.mileage_available is not None else "",
                raw.miles.miles_earned or "", raw.miles.bonus_percent if raw.miles.bonus_percent is not None else "",
                scrape_time, rec.get("status", ""),
            ]
            collected.append((row, amap, tail))

    amen_keys = used_amenity_keys([a for _r, a, _t in collected])
    amen_cols = [AMENITY_TR[k] for k in amen_keys]
    headers = BASE_COLS + amen_cols + TAIL_COLS
    ws.append(headers)
    for row, amap, tail in collected:
        cells = list(row)
        for k in amen_keys:
            state, det = amap[k]
            canon = canonical_rule_detail(k, AmenityStatus(state))
            if canon is not None:
                det = canon              # one standard vocabulary for rule rights
            cells.append(f"{state} — {det}" if (state != "Unknown" and det) else state)
        ws.append(cells + tail)
        # colour amenity cells
        r_idx = ws.max_row
        for j, k in enumerate(amen_keys):
            state = amap[k][0]
            cell = ws.cell(row=r_idx, column=len(BASE_COLS) + 1 + j)
            cell.fill = STATE_FILL[state]
            cell.font = STATE_FONT[state]
            cell.alignment = Alignment(horizontal="center")
    n = len(collected)

    # header styling + layout
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    widths = {"Carrier": 8, "Airline": 20, "OND": 10, "Origin": 8, "Destination": 11, "Season": 8,
              "Departure": 11, "Return": 11, "Cabin": 10, "Tier": 7, "Brand Order": 6,
              "Raw Brand Name": 22, "Normalized Brand": 20, "Fare Family Code": 16,
              "Display Price": 14, "Abs Price": 13, "Currency": 7, "Source": 10}
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 13 if h in amen_cols else 12)
    ws.row_dimensions[1].height = 34
    # price number format
    price_col = headers.index("Abs Price") + 1
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=price_col).number_format = "#,##0.00"

    wb.save(out_xlsx)
    dropped = len(AMENITY_KEYS) - len(amen_keys)
    print(f"Wrote {n} rows -> {out_xlsx} ({out_xlsx.stat().st_size//1024} KB)"
          f"{f' — {dropped} empty right column(s) dropped' if dropped else ''}")


if __name__ == "__main__":
    main()

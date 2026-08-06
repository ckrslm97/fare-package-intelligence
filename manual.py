#!/usr/bin/env python3
"""Manual package entry — add, correct and remove branded fares by hand.

Why this exists: the scrape reads a public site, so it can miss a package, get
a right wrong, or not cover a carrier at all. This lets a human state the truth
for any (carrier, OND, season, cabin) without editing the scraper's output by
hand and without a re-scrape wiping the correction.

THE ONE THING THAT MAKES THIS PRACTICAL
---------------------------------------
A package's RIGHTS come from its fare rules, so they belong to the
(carrier, brand) pair — "TK Flex Fly" carries the same rights on every route.
Only the PRICE changes per OND and season. Typing 16 rights for every route
would be thousands of cells; typing them ONCE per brand and then only a price
per route is a few dozen. So every right column may be left BLANK, and it is
inherited from:

  1. another manual row for the same (carrier, brand),
  2. failing that, the scraped dataset for the same (carrier, brand).

Fill a right only when you are correcting or introducing it.

MERGE RULE (chosen by the operator, 2026-08-05)
-----------------------------------------------
Package level. A manual row REPLACES the scraped package with the same brand
name and ADDS it when there is none; scraped packages you do not name are left
alone. `Action=remove` deletes a scraped package that should not be there.

USAGE
    python manual.py template                 # boş şablon üret (xlsx + csv)
    python manual.py export <out_dir> [--carrier TK] [--ond IST-LHR] [--season Summer]
    python manual.py add                      # tek tek, soru-cevap
    python manual.py import <dosya.csv|xlsx>  # toplu yükle (doğrular, birleştirir)
    python manual.py validate
    python manual.py apply <out_dir> [--out <hedef_dir>]
"""
from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from branded_fare_scraper.amenities import AMENITY_DISPLAY, AMENITY_KEYS
from branded_fare_scraper.models import AmenityStatus, Cabin, Season
from branded_fare_scraper.normalization import brand_match_key
from branded_fare_scraper.rebuild import iter_raw_records

STORE = Path("manual_data/entries.csv")

UNIT_COLS = ["Action", "Carrier", "Origin", "Destination", "Season", "Cabin"]
PKG_COLS = ["Brand Order", "Brand Name", "Price USD", "Step USD"]
RIGHT_COLS = [AMENITY_DISPLAY[k] for k in AMENITY_KEYS]
TAIL_COLS = ["Note"]
COLUMNS = UNIT_COLS + PKG_COLS + RIGHT_COLS + TAIL_COLS

STATUSES = {s.value.lower(): s.value for s in AmenityStatus}
STATUSES.update({"dahil": "Included", "var": "Included", "ucretli": "Paid",
                 "ücretli": "Paid", "yok": "Not Included", "-": "Not Included"})
SEASONS = {s.value.lower(): s.value for s in Season}
CABINS = {c.value.lower(): c.value for c in Cabin}
RIGHT_BY_DISPLAY = {AMENITY_DISPLAY[k]: k for k in AMENITY_KEYS}


# --------------------------------------------------------------------------- #
# store
# --------------------------------------------------------------------------- #
def read_store() -> list[dict]:
    if not STORE.exists():
        return []
    with STORE.open(encoding="utf-8-sig", newline="") as f:
        return [r for r in csv.DictReader(f) if any((v or "").strip() for v in r.values())]


def write_store(rows: list[dict]) -> None:
    STORE.parent.mkdir(parents=True, exist_ok=True)
    with STORE.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in COLUMNS})


def read_any(path: Path) -> list[dict]:
    """Read the bulk file — .csv or .xlsx, same columns either way."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl
        ws = openpyxl.load_workbook(path, read_only=True, data_only=True).active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        head = [str(c).strip() if c is not None else "" for c in rows[0]]
        out = []
        for r in rows[1:]:
            d = {head[i]: ("" if v is None else str(v).strip())
                 for i, v in enumerate(r) if i < len(head) and head[i]}
            if any(d.values()):
                out.append(d)
        return out
    with path.open(encoding="utf-8-sig", newline="") as f:
        return [r for r in csv.DictReader(f) if any((v or "").strip() for v in r.values())]


# --------------------------------------------------------------------------- #
# validation
# --------------------------------------------------------------------------- #
def norm_row(r: dict, lineno: int) -> tuple[dict | None, list[str]]:
    """Normalize one row; return (row, errors). Never guesses a unit key."""
    e: list[str] = []
    g = lambda c: (r.get(c) or "").strip()  # noqa: E731

    out = {c: "" for c in COLUMNS}
    out["Action"] = (g("Action") or "upsert").lower()
    if out["Action"] not in ("upsert", "remove"):
        e.append(f"Action '{out['Action']}' geçersiz (upsert veya remove)")

    for c, n in (("Carrier", 2), ("Origin", 3), ("Destination", 3)):
        v = g(c).upper()
        if not v:
            e.append(f"{c} boş")
        elif c == "Carrier" and not (2 <= len(v) <= 3):
            e.append(f"Carrier '{v}' 2-3 karakter olmalı")
        elif c != "Carrier" and not (len(v) == 3 and v.isalpha()):
            e.append(f"{c} '{v}' 3 harfli havalimanı kodu olmalı")
        out[c] = v

    s = SEASONS.get(g("Season").lower())
    if not s:
        e.append(f"Season '{g('Season')}' geçersiz ({'/'.join(x.value for x in Season)})")
    out["Season"] = s or ""

    cab = CABINS.get(g("Cabin").lower())
    if not cab:
        e.append(f"Cabin '{g('Cabin')}' geçersiz ({'/'.join(x.value for x in Cabin)})")
    out["Cabin"] = cab or ""

    out["Brand Name"] = g("Brand Name")
    if not out["Brand Name"]:
        e.append("Brand Name boş")

    for c in ("Brand Order", "Price USD", "Step USD"):
        v = g(c).replace(",", ".")
        if v:
            try:
                float(v)
            except ValueError:
                e.append(f"{c} '{v}' sayı değil")
            else:
                out[c] = v
    if out["Action"] == "upsert" and not (out["Price USD"] or out["Step USD"]):
        e.append("Price USD veya Step USD birinden biri gerekli")
    if out["Price USD"] and out["Step USD"]:
        e.append("Price USD ve Step USD birlikte verilemez")

    for c in RIGHT_COLS:
        v = g(c)
        if not v:
            continue
        head, _, detail = v.partition(":")
        st = STATUSES.get(head.strip().lower())
        if not st:
            e.append(f"{c} '{v}' geçersiz (Included / Paid / Not Included)")
        else:
            out[c] = f"{st}: {detail.strip()}" if detail.strip() else st
    out["Note"] = g("Note")
    return (None, [f"satır {lineno}: {x}" for x in e]) if e else (out, [])


def validate(rows: list[dict]) -> tuple[list[dict], list[str]]:
    ok, errs = [], []
    for i, r in enumerate(rows, start=2):
        row, e = norm_row(r, i)
        errs += e
        if row:
            ok.append(row)
    seen = {}
    for i, r in enumerate(ok, start=2):
        k = (r["Carrier"], r["Origin"], r["Destination"], r["Season"], r["Cabin"],
             brand_match_key(r["Brand Name"]))
        if k in seen:
            errs.append(f"satır {i}: aynı paket {seen[k]}. satırda da var ({r['Brand Name']})")
        seen[k] = i
    return ok, errs


# --------------------------------------------------------------------------- #
# rights inheritance
# --------------------------------------------------------------------------- #
def scraped_rights(raw_path: Path) -> dict[tuple, dict[str, str]]:
    """(carrier, brand-key) -> {right display name: "Status: detail"} from a scrape."""
    out: dict[tuple, dict[str, str]] = {}
    if not raw_path.exists():
        return out
    for rec in iter_raw_records(raw_path):
        car = rec.get("carrier", "")
        for c in rec.get("cabins", []):
            for b in c.get("brands", []):
                key = (car, brand_match_key(b.get("raw_brand_name", "")))
                slot = out.setdefault(key, {})
                for a in b.get("amenities", []):
                    disp = AMENITY_DISPLAY.get(a.get("canonical_key") or "")
                    if not disp or disp in slot:
                        continue
                    # Detayı da taşı: "Included" ile "Included: 2 parça X 8 kg"
                    # aynı şey değil, ve ağırlığı düşürmek tam olarak bu
                    # projenin tekrar tekrar avladığı hata sınıfı.
                    det = (a.get("raw_value") or "").strip()
                    st = a.get("status", "")
                    slot[disp] = f"{st}: {det}" if det else st
    return out


def resolve_rights(rows: list[dict], inherited: dict[tuple, dict[str, str]]) -> None:
    """Fill blank right cells in place, manual rows first, then the scrape."""
    from_manual: dict[tuple, dict[str, str]] = {}
    for r in rows:
        key = (r["Carrier"], brand_match_key(r["Brand Name"]))
        slot = from_manual.setdefault(key, {})
        for c in RIGHT_COLS:
            if r.get(c):
                slot.setdefault(c, r[c])
    for r in rows:
        key = (r["Carrier"], brand_match_key(r["Brand Name"]))
        for c in RIGHT_COLS:
            if not r.get(c):
                r[c] = from_manual.get(key, {}).get(c) or inherited.get(key, {}).get(c, "")


# --------------------------------------------------------------------------- #
# apply
# --------------------------------------------------------------------------- #
def to_amenities(row: dict) -> list[dict]:
    out = []
    for disp in RIGHT_COLS:
        v = (row.get(disp) or "").strip()
        if not v:
            continue
        st, _, detail = v.partition(":")
        out.append({"raw_label": RIGHT_BY_DISPLAY[disp], "status": st.strip(),
                    "raw_value": detail.strip(), "fee_amount": None,
                    "fee_currency": None, "canonical_key": RIGHT_BY_DISPLAY[disp]})
    return out


def apply(src: Path, dst: Path) -> None:
    rows, errs = validate(read_store())
    if errs:
        for e in errs[:20]:
            print(f"  ! {e}")
        raise SystemExit(f"{len(errs)} hata var; önce 'manual.py validate' ile düzelt.")
    if not rows:
        raise SystemExit("manual_data/entries.csv boş — uygulanacak giriş yok.")

    raw = src / "raw_data.jsonl"
    resolve_rights(rows, scraped_rights(raw))

    by_unit: dict[tuple, list[dict]] = {}
    for r in rows:
        by_unit.setdefault((r["Carrier"], r["Origin"], r["Destination"], r["Season"]), []).append(r)

    dst.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().isoformat(timespec="seconds")
    n_add = n_upd = n_del = n_new_unit = 0

    records: list[dict] = []
    seen_units: set[tuple] = set()
    for rec in iter_raw_records(raw):
        key = (rec.get("carrier"), rec.get("origin"), rec.get("destination"), rec.get("season"))
        seen_units.add(key)
        if key in by_unit:
            a, u, d = merge_unit(rec, by_unit[key], stamp)
            n_add += a; n_upd += u; n_del += d
        records.append(rec)

    for key, rs in by_unit.items():                 # units the scrape never had
        if key in seen_units:
            continue
        n_new_unit += 1
        rec = {"unit_key": f"manual::{'|'.join(key)}", "carrier": key[0], "origin": key[1],
               "destination": key[2], "season": key[3], "source": "manual",
               "status": "success", "retry_count": 0, "error": "", "elapsed_s": 0.0,
               "cabins": []}
        a, u, d = merge_unit(rec, rs, stamp)
        n_add += a; n_upd += u; n_del += d
        records.append(rec)

    with (dst / "raw_data.jsonl").open("w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False, default=str) + "\n")
    for extra in ("state",):                        # keep the frozen plan if present
        if (src / extra).exists() and not (dst / extra).exists():
            shutil.copytree(src / extra, dst / extra)

    bare = [r for r in rows if r["Action"] == "upsert"
            and not any(r.get(c) for c in RIGHT_COLS)]
    if bare:
        # Devralma marka bazlıdır: yepyeni bir marka adının devralacağı bir
        # geçmişi yoktur, o yüzden hakları elle yazılmalı. Sessiz kalırsa
        # panelde bomboş bir paket olarak görünür.
        print(f"\n  [uyarı] {len(bare)} pakette hiç hak yok — yeni bir marka adı,"
              f" devralacak kayıt bulunamadı:")
        for r in bare[:8]:
            print(f"      {r['Carrier']} {r['Origin']}-{r['Destination']} "
                  f"{r['Season']} {r['Cabin']} · {r['Brand Name']}")
        print("      (hakları şablonda doldurup tekrar import edin)\n")

    print(f"{src} -> {dst}")
    print(f"  eklenen paket   : {n_add}")
    print(f"  güncellenen     : {n_upd}")
    print(f"  silinen         : {n_del}")
    print(f"  yeni birim      : {n_new_unit}")
    print(f"  toplam ham kayıt: {len(records)}")
    print(f"\nşimdi:  ./.venv/bin/python reprocess_raw.py {dst}")


def merge_unit(rec: dict, rows: list[dict], stamp: str) -> tuple[int, int, int]:
    """Apply this unit's manual rows to one raw record, in place."""
    added = updated = removed = 0
    by_cabin: dict[str, list[dict]] = {}
    for r in rows:
        by_cabin.setdefault(r["Cabin"], []).append(r)

    for cabin, rs in by_cabin.items():
        cab = next((c for c in rec["cabins"] if c.get("cabin") == cabin), None)
        if cab is None:
            cab = {"cabin": cabin, "source": "manual", "departure": None, "return": None,
                   "has_availability": True, "note": f"manual {stamp}", "flight_no": "",
                   "booking_class": "", "operating_carrier": "", "is_codeshare": None,
                   "is_interline": None, "brands": []}
            rec["cabins"].append(cab)
        idx = {brand_match_key(b.get("raw_brand_name", "")): b for b in cab["brands"]}

        for r in rs:
            k = brand_match_key(r["Brand Name"])
            if r["Action"] == "remove":
                if k in idx:
                    cab["brands"].remove(idx.pop(k))
                    removed += 1
                continue
            brand = {
                "raw_brand_name": r["Brand Name"], "cabin": cabin,
                "screen_order": int(float(r["Brand Order"])) - 1 if r["Brand Order"] else 0,
                "price_value": float(r["Price USD"]) if r["Price USD"] else None,
                "price_type": "absolute", "currency": "USD",
                "display_price_text": "", "fare_family_code": None,
                "description": r.get("Note", ""), "amenities": to_amenities(r),
                "miles": {"mileage_available": None, "miles_earned": None,
                          "bonus_percent": None},
                "source": "manual",
                # carried so a later reader can tell a typed package from a
                # scraped one without diffing two files
                "_manual": {"at": stamp, "step_usd": r["Step USD"] or None},
            }
            if k in idx:
                cab["brands"][cab["brands"].index(idx[k])] = brand
                updated += 1
            else:
                cab["brands"].append(brand)
                added += 1
            idx[k] = brand

        # A "Step USD" row states the fee to reach this tier, so the absolute
        # price only exists once the ladder is ordered. Resolve after merging.
        cab["brands"].sort(key=lambda b: (b.get("screen_order") or 0,
                                          b.get("price_value") if b.get("price_value") is not None
                                          else float("inf")))
        running = None
        for i, b in enumerate(cab["brands"]):
            step = (b.get("_manual") or {}).get("step_usd")
            if step and running is not None:
                b["price_value"] = round(running + float(step), 2)
            if b.get("price_value") is not None:
                running = b["price_value"]
            b["screen_order"] = i
        if cab["brands"]:
            cab["has_availability"] = True
    rec["cabins"] = [c for c in rec["cabins"] if c["brands"]]
    return added, updated, removed


# --------------------------------------------------------------------------- #
# template / export / add
# --------------------------------------------------------------------------- #
LEGEND = [
    ["# FPI manuel giriş şablonu — doldurup 'python manual.py import <dosya>' ile yükleyin"],
    ["# Action    : upsert (ekle/değiştir, varsayılan) veya remove (çekilmiş paketi sil)"],
    ["# Season    : " + " / ".join(s.value for s in Season)],
    ["# Cabin     : " + " / ".join(c.value for c in Cabin)],
    ["# Fiyat     : 'Price USD' mutlak fiyat VEYA 'Step USD' bir önceki kademeden fark — biri"],
    ["# Haklar    : Included / Paid / 'Not Included'. Detay için iki nokta: 'Included: 2×23kg'"],
    ["# ÖNEMLİ    : Hak hücresini BOŞ bırakırsanız aynı taşıyıcı+marka için daha önce"],
    ["#             girdiğiniz değerden, o da yoksa çekilmiş veriden devralınır. Yani bir"],
    ["#             markanın haklarını BİR KEZ yazın; diğer hatlarda sadece fiyat girin."],
    [],
]


def cmd_template() -> None:
    Path("manual_data").mkdir(exist_ok=True)
    csv_p = Path("manual_data/sablon.csv")
    with csv_p.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        for line in LEGEND:
            w.writerow(line)
        w.writerow(COLUMNS)
        w.writerow(["upsert", "TK", "IST", "LHR", "Summer", "Economy", 1, "Eco Fly", 0, "",
                    "Included: 1×8kg", "Not Included", "", "Included"]
                   + [""] * (len(RIGHT_COLS) - 4) + ["örnek satır — silin"])
        w.writerow(["upsert", "TK", "IST", "LHR", "Summer", "Economy", 2, "Extra Fly", "", 29,
                    "", "Included: 1×23kg", "", ""]
                   + [""] * (len(RIGHT_COLS) - 4) + ["haklar boş = Eco Fly'dan devralınır"])
    try:
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Manuel Giris"
        for line in LEGEND:
            ws.append(line)
        ws.append(COLUMNS)
        for c in ws[len(LEGEND) + 1]:
            c.font = openpyxl.styles.Font(bold=True)
        ws.freeze_panes = ws.cell(row=len(LEGEND) + 2, column=1)
        for i, name in enumerate(COLUMNS, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(11, len(name) + 3)
        wb.save("manual_data/sablon.xlsx")
        print("manual_data/sablon.xlsx  ve  manual_data/sablon.csv yazıldı")
    except ImportError:
        print(f"{csv_p} yazıldı (openpyxl yok, xlsx atlandı)")


def cmd_export(out_dir: Path, carrier: str, ond: str, season: str,
               missing: str = "", cabin: str = "") -> None:
    """Dump existing packages in template shape so they are corrected, not retyped.

    ``missing`` narrows to the units where a right is absent everywhere — the
    shape a qa_check right-coverage warning takes, so closing one is: export
    what it flagged, fill the column, import.
    """
    raw = out_dir / "raw_data.jsonl"
    o = d = ""
    if ond:
        o, _, d = ond.upper().partition("-")
    if missing and missing not in RIGHT_COLS:
        raise SystemExit(f"--missing '{missing}' bilinmiyor. Seçenekler: {', '.join(RIGHT_COLS)}")
    rows = []
    for rec in iter_raw_records(raw):
        if carrier and rec.get("carrier", "").upper() != carrier.upper():
            continue
        if o and (rec.get("origin") != o or rec.get("destination") != d):
            continue
        if season and (rec.get("season", "").lower() != season.lower()):
            continue
        for c in rec.get("cabins", []):
            if cabin and c.get("cabin", "").lower() != cabin.lower():
                continue
            if missing:
                key = RIGHT_BY_DISPLAY[missing]
                has = any(a.get("canonical_key") == key
                          for b in c.get("brands", []) for a in b.get("amenities", []))
                if has:
                    continue          # bu birimde hak zaten var -> düzeltme gerekmiyor
            for i, b in enumerate(c.get("brands", []), start=1):
                r = {k: "" for k in COLUMNS}
                r.update({"Action": "upsert", "Carrier": rec["carrier"],
                          "Origin": rec["origin"], "Destination": rec["destination"],
                          "Season": rec["season"], "Cabin": c["cabin"],
                          "Brand Order": i, "Brand Name": b.get("raw_brand_name", ""),
                          "Price USD": b.get("price_value") or ""})
                for a in b.get("amenities", []):
                    disp = AMENITY_DISPLAY.get(a.get("canonical_key") or "")
                    if not disp:
                        continue
                    det = (a.get("raw_value") or "").strip()
                    r[disp] = f"{a.get('status')}: {det}" if det else a.get("status", "")
                rows.append(r)
    if not rows:
        raise SystemExit("filtreye uyan kayıt yok.")
    Path("manual_data").mkdir(exist_ok=True)
    p = Path("manual_data/duzenle.csv")
    with p.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        w.writeheader(); w.writerows(rows)
    print(f"{p}: {len(rows)} paket yazıldı — düzeltip 'manual.py import {p}' ile yükleyin")


def ask(prompt: str, default: str = "", options: list[str] | None = None) -> str:
    while True:
        hint = f" [{default}]" if default else ""
        v = input(f"  {prompt}{hint}: ").strip() or default
        if not options or v.lower() in [o.lower() for o in options]:
            return v
        print(f"    seçenekler: {', '.join(options)}")


def cmd_add() -> None:
    print("Tek paket ekle (boş bırakılan hak = aynı markadan devralınır)\n")
    base = {c: "" for c in COLUMNS}
    base["Action"] = ask("Action (upsert/remove)", "upsert", ["upsert", "remove"])
    base["Carrier"] = ask("Carrier (örn. TK)").upper()
    base["Origin"] = ask("Origin (örn. IST)").upper()
    base["Destination"] = ask("Destination (örn. LHR)").upper()
    base["Season"] = ask("Season", "Summer", [s.value for s in Season])
    base["Cabin"] = ask("Cabin", "Economy", [c.value for c in Cabin])
    base["Brand Name"] = ask("Paket adı")
    if base["Action"] == "upsert":
        base["Brand Order"] = ask("Kademe sırası (1,2,3…)", "")
        base["Price USD"] = ask("Mutlak fiyat USD (boş bırakıp adım verebilirsiniz)", "")
        if not base["Price USD"]:
            base["Step USD"] = ask("Bir önceki kademeden fark USD", "")
        print("  haklar — boş geç = devral, örn: 'Included: 2×23kg'")
        for disp in RIGHT_COLS:
            v = input(f"    {disp}: ").strip()
            if v:
                base[disp] = v
    base["Note"] = ask("Not", "")
    row, errs = norm_row(base, 0)
    if errs:
        for e in errs:
            print(f"  ! {e}")
        raise SystemExit("kaydedilmedi.")
    rows = read_store(); rows.append(row); write_store(rows)
    print(f"\neklendi -> {STORE} ({len(rows)} giriş)")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("template")
    sub.add_parser("add")
    sub.add_parser("validate")
    e = sub.add_parser("export"); e.add_argument("out_dir")
    e.add_argument("--carrier", default=""); e.add_argument("--ond", default="")
    e.add_argument("--season", default="")
    e.add_argument("--cabin", default="", help="tek kabin (Economy/Business/…)")
    e.add_argument("--missing", default="",
                   help="yalnız bu hakkın HİÇ olmadığı birimler (örn. --missing Meal)")
    i = sub.add_parser("import"); i.add_argument("path")
    i.add_argument("--replace", action="store_true", help="mevcut girişleri sil, baştan yaz")
    a = sub.add_parser("apply"); a.add_argument("out_dir"); a.add_argument("--out", default="")
    args = ap.parse_args()

    if args.cmd == "template":
        cmd_template()
    elif args.cmd == "add":
        cmd_add()
    elif args.cmd == "export":
        cmd_export(Path(args.out_dir), args.carrier, args.ond, args.season,
                   args.missing, args.cabin)
    elif args.cmd == "validate":
        rows, errs = validate(read_store())
        for x in errs[:40]:
            print(f"  ! {x}")
        print(f"{len(rows)} geçerli giriş, {len(errs)} hata")
        sys.exit(1 if errs else 0)
    elif args.cmd == "import":
        incoming, errs = validate(read_any(Path(args.path)))
        if errs:
            for x in errs[:40]:
                print(f"  ! {x}")
            raise SystemExit(f"{len(errs)} hata — hiçbir şey yüklenmedi.")
        cur = [] if args.replace else read_store()
        idx = {(r["Carrier"], r["Origin"], r["Destination"], r["Season"], r["Cabin"],
                brand_match_key(r["Brand Name"])): i for i, r in enumerate(cur)}
        added = updated = 0
        for r in incoming:
            k = (r["Carrier"], r["Origin"], r["Destination"], r["Season"], r["Cabin"],
                 brand_match_key(r["Brand Name"]))
            if k in idx:
                cur[idx[k]] = r; updated += 1
            else:
                idx[k] = len(cur); cur.append(r); added += 1
        write_store(cur)
        print(f"{args.path}: {added} yeni, {updated} güncellendi -> {STORE} ({len(cur)} giriş)")
    elif args.cmd == "apply":
        src = Path(args.out_dir)
        dst = Path(args.out) if args.out else Path(str(src).rstrip("/") + "_manual")
        apply(src, dst)


if __name__ == "__main__":
    main()

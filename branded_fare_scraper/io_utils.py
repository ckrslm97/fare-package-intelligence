"""Input reading + output writing.

Input: an Excel (``.xlsx``) or ``.csv`` with (at least) a Carrier column plus
either Origin+Destination columns or a single combined OND column ("YVR-DEL");
matching is case-insensitive and a few common aliases are accepted. Any other
column (PAX, RANK_CARRIER, regions, countries, ...) is carried through unchanged
in ``Job.raw_row``; unnamed columns (e.g. an exported index) are dropped.

Outputs (all under the output dir):
* ``raw_data.jsonl``        — raw scrape result per unit
* ``normalized_data.csv``   — flat rows, one column per amenity (always written)
* ``normalized_data.xlsx``  — same, if openpyxl is available
* ``failed_jobs.csv``       — units that failed / had no availability
* ``summary.json``          — run summary

openpyxl is imported lazily so the engine has no hard third-party dependency for
CSV-only workflows.
"""

from __future__ import annotations

import csv
import json
import re
from dataclasses import asdict
from pathlib import Path
from typing import Any, Iterable

from .airports import meta as airport_meta
from .amenities import AMENITY_DISPLAY, AMENITY_KEYS
from .models import BrandedFare, Job, RunSummary, UnitResult
from .normalization import clean_fare_code

# --------------------------------------------------------------------------- #
# Input
# --------------------------------------------------------------------------- #
_CARRIER_ALIASES = {"carrier", "airline", "carriercode", "carrier_code", "crr", "taşıyıcı", "tasiyici"}
_ORIGIN_ALIASES = {"origin", "orgn", "from", "dep", "departure", "o", "kalkış", "kalkis"}
_DEST_ALIASES = {"destination", "dest", "to", "arrival", "arr", "d", "varış", "varis"}
_OND_ALIASES = {"ond", "route", "od", "o-d"}


def _match(colname: str, aliases: set[str]) -> bool:
    return colname.strip().lower().replace(" ", "") in aliases


def _cell(row: list[Any], i: int | None) -> str:
    """Value at column ``i`` as a stripped string ('' for missing/None)."""
    if i is None or i >= len(row):
        return ""
    v = row[i]
    return "" if v is None else str(v).strip()


def _split_ond(ond: str) -> tuple[str, str] | None:
    """Split 'LHR-SIN' (or 'LHR/SIN') into (origin, destination)."""
    parts = re.split(r"[-/_>]", str(ond).strip().upper())
    parts = [p for p in parts if p]
    if len(parts) == 2 and all(2 <= len(p) <= 4 for p in parts):
        return parts[0], parts[1]
    return None


def _rows_to_jobs(headers: list[str], rows: Iterable[list[Any]]) -> list[Job]:
    # Role detection is independent per column: a header is tested against every
    # alias set, so column order never hides a role. Unnamed columns (e.g. the
    # pandas index column exported into xlsx) take no role and are not carried
    # into raw_row.
    idx: dict[str, int | None] = {"carrier": None, "origin": None, "destination": None, "ond": None}
    _ROLES = (
        ("carrier", _CARRIER_ALIASES),
        ("origin", _ORIGIN_ALIASES),
        ("destination", _DEST_ALIASES),
        ("ond", _OND_ALIASES),
    )
    named: list[int] = []                     # column indices with a real header
    for i, h in enumerate(headers):
        h = str(h if h is not None else "").strip()
        if not h or h.lower() == "none":
            continue
        named.append(i)
        for role, aliases in _ROLES:
            if idx[role] is None and _match(h, aliases):
                idx[role] = i
                break

    if idx["carrier"] is None:
        raise ValueError(f"Input is missing a Carrier column. Headers seen: {headers}")
    has_od = idx["origin"] is not None and idx["destination"] is not None
    if not has_od and idx["ond"] is None:
        raise ValueError(f"Input needs either Origin+Destination or an OND column. Headers: {headers}")

    jobs: list[Job] = []
    seen: set[tuple[str, str, str]] = set()
    for row in rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        # every other column (PAX, RANK_CARRIER, regions, countries, ...) passes
        # through untouched
        raw_row = {str(headers[i]): row[i] for i in named if i < len(row)}
        carrier = _cell(row, idx["carrier"])
        if has_od and _cell(row, idx["origin"]) and _cell(row, idx["destination"]):
            origin = _cell(row, idx["origin"]).upper()
            destination = _cell(row, idx["destination"]).upper()
        elif idx["ond"] is not None and _cell(row, idx["ond"]):
            split = _split_ond(_cell(row, idx["ond"]))   # "YVR-DEL" -> ("YVR", "DEL")
            if not split:
                continue
            origin, destination = split
        else:
            continue
        if not (carrier and origin and destination):
            continue
        key = (carrier.upper(), origin, destination)
        if key in seen:                      # dedupe repeated (carrier, OND) rows
            continue
        seen.add(key)
        jobs.append(Job(carrier=carrier, origin=origin, destination=destination, raw_row=raw_row))
    return jobs


def read_jobs(path: str) -> list[Job]:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Input file not found: {p}")
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook  # lazy
        wb = load_workbook(p, read_only=True, data_only=True)
        try:
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
        if not data:
            return []
        headers = [str(c) if c is not None else "" for c in data[0]]
        return _rows_to_jobs(headers, [list(r) for r in data[1:]])
    # CSV
    with p.open("r", encoding="utf-8-sig", newline="") as f:
        reader = list(csv.reader(f))
    if not reader:
        return []
    return _rows_to_jobs(reader[0], reader[1:])


# --------------------------------------------------------------------------- #
# Output schema
# --------------------------------------------------------------------------- #
BASE_COLUMNS = [
    "Carrier", "Origin", "Destination",
    "Origin City Code", "Origin Country Code", "Dest City Code", "Dest Country Code",
    "Departure Date", "Return Date", "Season",
    "Cabin", "Brand Order", "Tier Code", "Raw Brand Name", "Normalized Brand Name",
    "Display Price", "Calculated Absolute Price", "Currency", "Source",
    "Fare Family Code", "Brand Description",
]
AMENITY_COLUMNS = [AMENITY_DISPLAY[k] for k in AMENITY_KEYS]
TAIL_COLUMNS = [
    "Mileage Available", "Miles Earned", "Bonus Percent",
    "Scrape Timestamp", "Status", "Retry Count",
]
ALL_COLUMNS = BASE_COLUMNS + AMENITY_COLUMNS + TAIL_COLUMNS


def flatten_branded_fare(bf: BrandedFare) -> dict[str, Any]:
    o_meta = airport_meta(bf.origin)
    d_meta = airport_meta(bf.destination)
    row: dict[str, Any] = {
        "Carrier": bf.carrier,
        "Origin": bf.origin,
        "Destination": bf.destination,
        "Origin City Code": o_meta["city_code"],
        "Origin Country Code": o_meta["country_code"],
        "Dest City Code": d_meta["city_code"],
        "Dest Country Code": d_meta["country_code"],
        "Departure Date": bf.departure_date.isoformat() if bf.departure_date else "",
        "Return Date": bf.return_date.isoformat() if bf.return_date else "",
        "Season": bf.season.value,
        "Cabin": bf.cabin.value,
        "Brand Order": bf.brand_order,
        "Tier Code": bf.tier_code,
        "Raw Brand Name": bf.raw_brand_name,
        "Normalized Brand Name": bf.normalized_brand_name,
        "Display Price": bf.display_price,
        "Calculated Absolute Price": bf.calculated_absolute_price,
        "Currency": bf.currency or "",
        "Source": bf.source,
        "Fare Family Code": clean_fare_code(bf.fare_family_code),
        "Brand Description": bf.brand_description,
        "Mileage Available": ("Yes" if bf.mileage_available else "No") if bf.mileage_available is not None else "",
        "Miles Earned": bf.miles_earned if bf.miles_earned is not None else "",
        "Bonus Percent": bf.bonus_percent if bf.bonus_percent is not None else "",
        "Scrape Timestamp": bf.scrape_timestamp,
        "Status": bf.status,
        "Retry Count": bf.retry_count,
    }
    for key in AMENITY_KEYS:
        row[AMENITY_DISPLAY[key]] = bf.amenities.get(key, "Unknown")
    return row


def write_normalized(rows: list[BrandedFare], out_dir: Path) -> tuple[Path, Path | None]:
    flat = [flatten_branded_fare(r) for r in rows]
    csv_path = out_dir / "normalized_data.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=ALL_COLUMNS)
        w.writeheader()
        w.writerows(flat)

    xlsx_path: Path | None = None
    try:
        from openpyxl import Workbook  # lazy
        wb = Workbook()
        ws = wb.active
        ws.title = "Normalized"
        ws.append(ALL_COLUMNS)
        for row in flat:
            ws.append([row.get(c, "") for c in ALL_COLUMNS])
        ws.freeze_panes = "A2"
        xlsx_path = out_dir / "normalized_data.xlsx"
        wb.save(xlsx_path)
    except ImportError:
        xlsx_path = None                       # openpyxl not installed -> CSV is enough
    except Exception as e:  # noqa: BLE001 - real write error: warn, keep CSV
        import logging
        logging.getLogger("bfs").warning("xlsx write failed (%s); CSV written", e)
        xlsx_path = None
    return csv_path, xlsx_path


def unit_record(u: UnitResult) -> dict[str, Any]:
    """One raw JSONL record for a finished unit."""
    return {
        "unit_key": u.unit.key(),
        "carrier": u.unit.job.carrier,
        "origin": u.unit.job.origin,
        "destination": u.unit.job.destination,
        "season": u.unit.date_plan.season.value,
        "source": u.source,
        "status": u.status.value,
        "retry_count": u.retry_count,
        "error": u.error,
        "elapsed_s": u.elapsed,
        "cabins": [
            {
                "cabin": c.cabin.value,
                "source": c.source,
                "departure": c.departure.isoformat() if c.departure else None,
                "return": c.return_date.isoformat() if c.return_date else None,
                "has_availability": c.has_availability,
                "note": c.note,
                # Itinerary identity: written even when empty so a later
                # reprocessing can tell "source did not say" from "not stored".
                "flight_no": c.flight_no,
                "booking_class": c.booking_class,
                "operating_carrier": c.operating_carrier,
                "is_codeshare": c.is_codeshare,
                "is_interline": c.is_interline,
                "brands": [_raw_brand_to_dict(b) for b in c.brands],
            }
            for c in u.cabin_results
        ],
    }


def append_raw(unit: UnitResult, out_dir: Path) -> Path:
    """Append ONE finished unit to raw_data.jsonl, immediately.

    A run that is stopped (or sleeps, or crashes) used to lose every hour of
    scraping because the raw file was only written at the end. Each unit is now
    durable the moment it finishes; ``write_raw`` later rewrites the file with
    the in-memory results merged over whatever is already on disk, so a resumed
    run never double-writes a ``unit_key``.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "raw_data.jsonl"
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(unit_record(unit), ensure_ascii=False, default=str) + "\n")
    return path


def write_raw(units: list[UnitResult], out_dir: Path) -> Path:
    """Rewrite raw_data.jsonl = what is on disk, overlaid by this run's units.

    Deduped on ``unit_key`` (this run wins), so the incremental appends above
    and a resumed run's earlier records collapse into exactly one record each.
    """
    path = out_dir / "raw_data.jsonl"
    records: dict[str, dict] = {}
    order: list[str] = []
    if path.exists():                       # keep earlier/interrupted work
        with path.open(encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except json.JSONDecodeError:
                    continue                # a half-written line from a hard kill
                key = rec.get("unit_key") or f"__anon_{len(order)}"
                if key not in records:
                    order.append(key)
                records[key] = rec
    for u in units:
        rec = unit_record(u)
        key = rec["unit_key"]
        if key not in records:
            order.append(key)
        records[key] = rec
    with path.open("w", encoding="utf-8") as f:
        for key in order:
            f.write(json.dumps(records[key], ensure_ascii=False, default=str) + "\n")
    return path


def _raw_brand_to_dict(b) -> dict[str, Any]:
    d = asdict(b)
    # enums -> values
    d["cabin"] = b.cabin.value
    d["price_type"] = b.price_type.value
    for a in d.get("amenities", []):
        a["status"] = a["status"].value if hasattr(a["status"], "value") else a["status"]
    return d


def write_failed(units: list[UnitResult], out_dir: Path) -> Path:
    path = out_dir / "failed_jobs.csv"
    cols = ["Carrier", "Origin", "Destination", "Season", "Status", "Retry Count",
            "Sources Tried", "Error", "Validation Issues", "Elapsed (s)"]
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for u in units:
            w.writerow({
                "Carrier": u.unit.job.carrier,
                "Origin": u.unit.job.origin,
                "Destination": u.unit.job.destination,
                "Season": u.unit.date_plan.season.value,
                "Status": u.status.value,
                "Retry Count": u.retry_count,
                "Sources Tried": ", ".join(u.sources_tried),
                "Error": u.error,
                "Validation Issues": "; ".join(u.validation_issues),
                "Elapsed (s)": u.elapsed,
            })
    return path


def write_summary(summary: RunSummary, out_dir: Path) -> Path:
    path = out_dir / "summary.json"
    path.write_text(json.dumps(summary.to_dict(), ensure_ascii=False, indent=2), "utf-8")
    return path
